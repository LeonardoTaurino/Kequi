from tkinter import *
import math
import ast
import openpyxl as xl
from openpyxl.chart import *
import os

path=filePath=os.path.abspath(os.path.dirname(__file__)) + "\\" #find script's directory path
fileResults="RISULTATI.txt" #iterazion results
fileExcel="Kequi.xlsx" # file excel, draw Kequi vs T and DG vs T
fileParameters="TERMODINAMICA.txt" #substance's parameters

class Point: #define a class which contein temperature , kp and GRT found during the iteration
    def __init__(self, x, y, z):
        self.x = x
        self.y = y
        self.z = z
    def disegno(self):
        #Save Results in a file txt
        file = open(path + fileResults, 'w')
        for i in range(len(self.x)):
            file.write(f"{self.x[i]},{self.y[i]},{self.z[i]}\n")
        file.close()
        #Save results in a file excel and draw graph
        wb = xl.load_workbook('Kequi.xlsx')
        sheet = wb['Foglio1']
        sheet.delete_cols(1)
        sheet.delete_cols(2)
        sheet.delete_cols(3)
        sheet.cell(1,1).value= 'Temperatura [K]'
        sheet.cell(1,2).value= 'Kequi'
        sheet.cell(1,3).value= 'DG°R(T) [J/mol]'
        for row in range(len(self.x)):
            cella_temperatura = sheet.cell(row+2, 1)
            cella_temperatura.value = self.x[row]
            cella_kequi = sheet.cell(row+2, 2)
            cella_kequi.value = self.y[row]
            cella_GT = sheet.cell(row+2, 3)
            cella_GT.value = self.z[row]
        chart1 = ScatterChart()
        chart1.title = "Kequi(T)"
        chart1.style = 6
        chart1.x_axis.title = 'Temperatura [K]'
        chart1.y_axis.title = 'Kequi'
        xvalue = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=1, max_col=1)
        yvalue1 = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=2, max_col=2)
        chart1.y_axis.crosses = "max"
        serie1 = Series(values = yvalue1, xvalues = xvalue)
        chart1.series.append(serie1)
        sheet.add_chart(chart1, "F2")
        chart2 = ScatterChart()
        chart2.title = "DG°R(T)"
        chart2.style = 6
        chart2.x_axis.title = 'Temperatura [K]'
        chart2.y_axis.title = 'DG°R(T) [J/mol]'
        yvalue2 = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=3, max_col=3)
        serie2 = Series(values=yvalue2, xvalues=xvalue)
        chart2.series.append(serie2)
        sheet.add_chart(chart2, "F18")
        wb.save('Kequi.xlsx')

#------------------------------------------------------------------------------------------------------------------
def check_switch(bottone, AA, BB, CC, DD, EE, stt): #Enable/Disable Iteration calculation
    stato=stt.get()
    if stato==0:
        bottone["state"] = NORMAL
        DD["state"] = NORMAL
        AA["state"] = DISABLED
        BB["state"] = DISABLED
        CC["state"] = DISABLED
        EE["state"] = DISABLED
    elif stato==1:
        bottone["state"] = DISABLED
        DD["state"] = DISABLED
        AA["state"] = NORMAL
        BB["state"] = NORMAL
        CC["state"] = NORMAL
        EE["state"] = NORMAL
#---------------------------------------------------------------------------------------------------------------------
def switch(bottone, stato): #Enable/Disable button/entry
    if stato=="on":
        bottone["state"] = NORMAL
    elif stato=="off":
        bottone["state"] = DISABLED
#---------------------------------------------------------------------------------------------------------------------
def Numero_sostanze(entryget): #Save number of substaces and use it to build window data and matrix data
    global num, datx
    try:
        num=int(entryget)
        datx=[[0 for x in range(10)] for y in range(num)]
        win.destroy()
    except ValueError:
        error4 = Toplevel()
        error4.title("ERROR")
        lbl_error3 = Label(error4, text="Enter the NUMBER of substances.").grid(row=0, column=0)
        btn_error3 = Button(error4, text="Ok", command=error4.destroy, padx=15, pady=5).grid(row=2, column=0)
#---------------------------------------------------------------------------------------------------------------------
def Costruzione_tabella(): #Build row by row data's window
    i=0
    for i in range(num):
        datx[i][9] = Entry(root, width=15)
        datx[i][9].grid(row=5+i, column=3)
        datx[i][8] = Button(root, text="Search", command=lambda i=i: Search(datx[i][9], i), padx=15)
        datx[i][8].grid(row=5 + i, column=2)
        #----------------------------------^^^^^
        datx[i][0]=Entry(root, width=15)
        datx[i][0].grid(row=5+i, column=4)
        datx[i][1]=Entry(root, width=15)
        datx[i][1].grid(row=5+i, column=5)
        datx[i][2]=Entry(root, width=15)
        datx[i][2].grid(row=5+i, column=6)
        datx[i][3]=Entry(root, width=15)
        datx[i][3].grid(row=5+i, column=7)
        datx[i][4]=Entry(root, width=15)
        datx[i][4].grid(row=5+i, column=8)
        datx[i][5]=Entry(root, width=15)
        datx[i][5].grid(row=5+i, column=9)
        datx[i][6]=Entry(root, width=15)
        datx[i][6].grid(row=5+i, column=10)
        datx[i][7]=Button(root, text="Save", command=lambda i=i: save(datx[i][9], i), padx=15)
        datx[i][7].grid(row=5+i, column=11)
#---------------------------------------------------------------------------------------------------------------------
def Search(specie, riga): #Search data from file txt and put them in the window's entry if previously saved
    spc=specie.get().upper()
    file = open(path + fileParameters)
    termo_str = file.read()
    termo = ast.literal_eval(termo_str)
    file.close()
    for a in termo.values():
        if spc in a:
            coeff=list(list(termo.keys())[list(termo.values()).index(a)])
            for m in range(6):
                datx[riga][m+1].delete(0, END)
                datx[riga][m+1].insert(0, coeff[m])
            switch(datx[riga][7], 'off')
#---------------------------------------------------------------------------------------------------------------------
def save(specie_s, riga_s): #save substances' parameter in the txt file with the given name if the substance is not saved already
    file = open(path + fileParameters)
    termo_str = file.read()
    termo = ast.literal_eval(termo_str)
    file.close()
    esist=0 # check if the substance exist in the database
    mm=0 # check variable
    spc=specie_s.get().upper()
    for a in termo.values(): #verify that the substance is not in the database
        if spc in a:
            mm+=1
            esist=1
    if spc == '':
        mm+=1
    for c in range(6):
        if datx[riga_s][c+1].get()=='':
            mm+=1
    if mm>0 and esist==0:
        error1 = Toplevel()
        error1.title("ERROR")
        lbl_error_win = Label(error1, text="Complete all data\nPlease write: DH°, DG°, C1, C2, C3, C4").grid(row=0, column=0)
        btn_error_win = Button(error1, text="Ok", command=error1.destroy, padx=15, pady=5).grid(row=2, column=0)
        #error1.mainloop()
    if esist>0:
        error1 = Toplevel()
        error1.title("ERROR")
        lbl_error_win = Label(error1, text=f"The substance {spc} is already in the database").grid(row=0, column=0)
        btn_error_win = Button(error1, text="Ok", command=error1.destroy, padx=15, pady=5).grid(row=2, column=0)
        #error1.mainloop()

    if mm==0 and esist==0:
        #costruisco il tuple da inserire nel dict
        vettor=[0,0,0,0,0,0]
        for y in range(6):
            vettor[y]=datx[riga_s][y+1].get()
        vet=tuple(vettor)
        termo.setdefault(vet, spc)
        file = open(path + fileParameters, 'w')
        file.write(str(termo))
        file.close()
        error1 = Toplevel()
        error1.title("SAVED!")
        lbl_error_win = Label(error1, text="Data saved with success").grid(row=0, column=0)
        btn_error_win = Button(error1, text="Ok", command=error1.destroy, padx=15, pady=5).grid(row=2, column=0)
        #error1.mainloop()

    del mm, esist, y
#---------------------------------------------------------------------------------------------------------------------
def Cph(tau, a, b, c, d): # calculate cph and cps
    b=b*10**(-3)
    c=c*10**(-6)
    d=d*10**(5)
    return ((a*tau+b*0.5*tau**2+(c*tau**3)/3-d/tau)-(a*298+b*0.5*298**2+(c*298**3)/3-d/298))*8.314
def Cps(tau, a, b, c, d):
    b = b * 10 ** (-3)
    c = c * 10 ** (-6)
    d = d * 10 ** (5)
    return ((a*math.log(tau)+b*tau+c*0.5*tau**2-0.5*d/(tau**2))-(a*math.log(298)+b*298+c*0.5*298**2-0.5*d/(298**2)))*8.314
#---------------------------------------------------------------------------------------------------------------------
def elaborat(dati, tau): #Save datas in lists and calculate
    i=0
    j=0
    try:
        T=float(tau.get())
        mela = [[0 for x in range(8)] for y in range(num)]

        for i in range(num):
            for j in range(7):
                mela[i][j]=float(dati[i][j].get())

        Grif=0
        Hrif=0
        CPH=0
        CPS=0
        i=0
        for i in range(num):
            Hrif+=mela[i][0]*mela[i][1]
            Grif+=mela[i][0]*mela[i][2]
            CPH+=mela[i][0]*Cph(T, mela[i][3], mela[i][4], mela[i][5], mela[i][6])
            CPS+=mela[i][0]*Cps(T, mela[i][3], mela[i][4], mela[i][5], mela[i][6])
        Srif=(Hrif-Grif)/298
        S=Srif+CPS
        H=Hrif+CPH
        G=H-T*S
        ex=(-G/(8.314*T))
        try:
            kp=math.exp(ex)
        except OverflowError:
            kp='inf'
        #Inserisco Output risultati finali

        risul = Toplevel()
        risul.title("RESULTS")
        lbl_risul1 = Label(risul, text=f"Kequi= {kp}").grid(row=0, column=0)
        lbl_risul2 = Label(risul, text=f"DG°R(T) [J/mol]= {G}").grid(row=1, column=0)
        lbl_risul3 = Label(risul, text=f"Dh°R(T) [J/mol]= {H}").grid(row=2, column=0)
        lbl_risul4 = Label(risul, text=f"DS°R(T) [J/mol*K]= {S}").grid(row=3, column=0)
        lbl_risul5 = Label(risul, text=f"DG°R(298) [J/mol]= {Grif}").grid(row=4, column=0)
        lbl_risul6 = Label(risul, text=f"DH°R(298) [J/mol]= {Hrif}").grid(row=5, column=0)
        lbl_risul7 = Label(risul, text=f"DS°R(T) [J/mol*K]= {Srif}").grid(row=6, column=0)
        lbl_risul8 = Label(risul, text=f"Cph [J/mol*K]= {CPH}").grid(row=7, column=0)
        lbl_risul9 = Label(risul, text=f"Cp2 [J/mol*K]= {CPS}").grid(row=8, column=0)
        btn_error_win = Button(risul, text="Ok", command=risul.destroy, padx=15, pady=5).grid(row=10, column=0)
        # error1.mainloop()
    except ValueError:
        error3 = Toplevel()
        error3.title("ERROR")
        lbl_error3 = Label(error3, text="Complete all field requested.").grid(row=0, column=0)
        btn_error3 = Button(error3, text="Ok", command=error3.destroy, padx=15, pady=5).grid(row=2, column=0)
#Definisco f per iterare i dati---------e salva i dati in RISULTATI.TXT--------------------------------------------------------------------
def iterazione(dati, min, max, passo):
    
    try:
        T1 = float(min.get())
        T1 = int(round(T1))
        T2 = float(max.get())
        T2 = int(round(T2))
        DT = float(passo.get())
        DT = int(round(DT))
        pera = [[0 for x in range(7)] for y in range(num)]
        for i in range(num):
            for j in range(7):
                pera[i][j] = float(dati[i][j].get())

        ST=[]
        HT=[]
        GT=[]
        KpT=[]
        i = 0
        for T in range(T1, T2+DT, DT):
            Grif = 0
            Hrif = 0
            CPH = 0
            CPS = 0
            for i in range(num):
                Hrif+=pera[i][0]*pera[i][1]
                Grif+=pera[i][0]*pera[i][2]
                CPH+=pera[i][0]*Cph(T, pera[i][3], pera[i][4], pera[i][5], pera[i][6])
                CPS+=pera[i][0]*Cps(T, pera[i][3], pera[i][4], pera[i][5], pera[i][6])

            Srif=(Hrif-Grif)/298
            S=Srif+CPS
            H=Hrif+CPH
            G=H-T*S
            ex=(-G/(8.314*T))
            try:
                kp=math.exp(ex)
            except OverflowError:
                kp='inf'
            ST.append(S)
            HT.append(H)
            GT.append(G)
            KpT.append(kp)
        temp = list(range(T1, T2+DT, DT))
        point = Point(temp, KpT, GT)
        point.disegno()
        risul = Toplevel()
        risul.title("RESULTS")
        lbl_risul1 = Label(risul, text="Iteration ended successfully!\nResults were saved in:\n" + path + fileResults).grid(row=0, column=0)
        btn_error_win = Button(risul, text="Ok", command=lambda:[risul.destroy(), os.system(path + fileExcel)], padx=15, pady=5).grid(row=10, column=0)
    except ValueError:
        error3 = Toplevel()
        error3.title("ERROR")
        lbl_error3 = Label(error3, text="Complete all field requested.").grid(row=0, column=0)
        btn_error3 = Button(error3, text="Ok", command=error3.destroy, padx=15, pady=5).grid(row=2, column=0)
    return


#------------------------------------------Finestra 1------------------------------------------
win = Tk()
win.title("Number of substances")
lbl_win1=Label(win, text="Enter the number of substances: ").grid(row=0, column=0)
e_win1=Entry(win, width=15)
e_win1.grid(row=0, column=1)
btn_win=Button(win, text="Ok", command=lambda:[Numero_sostanze(e_win1.get())], padx=15).grid(row=0, column=2)
win.mainloop()
#---------------------------------Nuova finestra--> finestra Dati------------------------------------

root = Tk()
root.title("Data")
lbl_root0=Label(root, text="Final Temperature [K]= ").grid(row=0, column=0)
e_root0=Entry(root, width=10)
e_root0.grid(row=0, column=1)
#-----------ITERAZIONE--------Creo label per iterazione con checkboxs che ingrigia CALCOLA e TEMPERATURA----------
lbl_rootA=Label(root, text="Temp,min [K]= ").grid(row=0, column=4)
e_rootA=Entry(root, state=DISABLED, width=15)
e_rootA.grid(row=0, column=5)
lbl_rootB=Label(root, text="Temp,max [K]= ").grid(row=0, column=6)
e_rootB=Entry(root, state=DISABLED, width=15)
e_rootB.grid(row=0, column=7)
lbl_rootC=Label(root, text="DeltaT step= ").grid(row=0, column=8)
e_rootC=Entry(root, state=DISABLED, width=15)
e_rootC.grid(row=0, column=9)
btn_rootA=Button(root, state=DISABLED, text="ITERA!", command=lambda:[iterazione(datx, e_rootA, e_rootB, e_rootC)], width=12)
btn_rootA.grid(row=0,column=10)
var=IntVar()
check=Checkbutton(root, text="Iteration", variable=var, command=lambda:[check_switch(btn_root0,e_rootA,e_rootB,e_rootC,e_root0,btn_rootA, var)])
check.grid(row=0, column=3)
#--------------Inserisco 2 righe dati sopra la griglia---------------------------------------------------------------
lbl_gost=Label(root, text="").grid(row=1, column=0)
lbl_root1=Label(root, text="Search", width=15)
lbl_root1.grid(row=3, column=2)
lbl_root2=Label(root, text="Substance", width=15)
lbl_root2.grid(row=3, column=3)
lbl_rootni=Label(root, text="stech Num.", width=15)
lbl_rootni.grid(row=3, column=4)
lbl_root3=Label(root, text="DH°f(298)", width=15)
lbl_root3.grid(row=3, column=5)
lbl_root4=Label(root, text="[Joule/mole]", width=15)
lbl_root4.grid(row=4, column=5)
lbl_root5=Label(root, text="DG°f(298)", width=15)
lbl_root5.grid(row=3, column=6)
lbl_root6=Label(root, text="[Joule/mole]", width=15)
lbl_root6.grid(row=4, column=6)
lbl_root7=Label(root, text="C1", width=15)
lbl_root7.grid(row=3, column=7)
lbl_root8=Label(root, text="C2", width=15)
lbl_root8.grid(row=3, column=8)
lbl_root9=Label(root, text="C3", width=15)
lbl_root9.grid(row=3, column=9)
lbl_root10=Label(root, text="C4", width=15)
lbl_root10.grid(row=3, column=10)
lbl_root11=Label(root, text="store data", width=15)
lbl_root11.grid(row=3, column=11)
#Costruisco la griglia dei dati------------------------------------------------------
Costruzione_tabella()
#Button to start calculation-------------------------------------------------------
btn_root0=Button(root, text="START!", command=lambda:[elaborat(datx, e_root0)])
btn_root0.grid(row=0,column=2)
#---------------------------------END----------------------------------------
root.mainloop()